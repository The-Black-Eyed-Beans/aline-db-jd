from os import listdir
from os.path import isfile, join
from logger import Logger
from openpyxl import load_workbook
from helpers import process_data, process_table

logger = Logger("data_dump.log")
logger.start_logger()

logger.add("DEBUG","Strating Aline database injection script...")

read_files = []
mypath = "."
fileList = [f for f in listdir(mypath) if isfile(join(mypath, f)) and (f.endswith(".xls") or f.endswith(".xlsx"))]

for file_ in fileList:
    record_list = {}
    if file_ in read_files:
        logger.add("ERROR","File: %s has already been processed. Skipping file.",file_)
        continue
    logger.add("DEBUG","Reading file: %s",file_)
    try:
        read_files.append(file_)
        wb = load_workbook(file_)
        logger.add("DEBUG","File successfully read.")
        if len(wb.sheetnames) != 0:
            logger.add("DEBUG","Sheets found!")
            logger.add("DEBUG","%s",wb.sheetnames)
    except:
        logger.add("ERROR","Error ocurred in loading file. Aborting file.")
        continue
    for sheet in wb.sheetnames:
        logger.add("DEBUG","Checking workbook for %s.",sheet)
        ws = wb[sheet]
        ws_len = len(ws['A'])
        if ws_len <= 1:
            logger.add("DEBUG","0 records found. Skipping table.")
            record_list[sheet] = []
            continue
        records = process_table(ws)
        record_list[sheet] = records 
    process_data(record_list)
    logger.add("DEBUG","Data processing complete!")
print("log @ data_dump.log")
