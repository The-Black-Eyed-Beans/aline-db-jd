from logger import Logger
from openpyxl.utils import get_column_letter
from conn import get_conn

logger = Logger()

def process_data(records):
    logger.add("DEBUG","Beginning record processing...")
    if len(records['bank']) != 0: add_data(records['bank'], 'bank')
    if len(records['branch']) != 0: add_data(records['branch'], 'branch')
    if len(records['applicant']) != 0: add_data(records['applicant'], 'applicant')
    if len(records['member']) != 0: add_data(records['member'], 'member')
    if len(records['user']) != 0: add_data(records['user'], 'user')
    if len(records['application']) != 0: add_data(records['application'], 'application')
    if len(records['merchant']) != 0: add_data(records['merchant'], 'merchant')
    if len(records['account']) != 0: add_data(records['account'], 'account')
    if len(records['transaction']) != 0: add_data(records['transaction'], 'transaction')

def add_data(records,tbl):
    logger.add("DEBUG","Table - %s",tbl)
    conn = get_conn()
    curs = conn.cursor()
    vals = ["%s"] * len(records[0])
    sql = "INSERT INTO %s VALUES (%s)" % (tbl, ",".join(vals))
    logger.add("DEBUG","Executing statement: %s",sql)
    try:
        curs.executemany(sql,records)
        conn.commit()
        logger.add("DEBUG","Execution complete. Records commited: %s",curs.rowcount)
    except:
        logger.add("ERROR","Execution failed. Skipping table!")

def process_table(ws):
    logger.add("DEBUG"," Records found. Processing...")
    col_len = len(tuple(ws.columns)) +1
    row_len = len(tuple(ws.rows)) +1
    records = []
    logger.add("DEBUG"," Total records found: %d",row_len-2)
    for row in range(2, row_len):
        record = []
        for col in range(1, col_len):
            char = get_column_letter(col)
            record.append(ws[char + str(row)].value)
        if len(record) != 0: records.append(tuple(record))
    logger.add("DEBUG"," Total records parsed: %d",len(records))
    return records
